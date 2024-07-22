import openpyxl
import re

def find_all_employees_data(sheet):
    employees_data = []
    current_employee = None

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[0] and isinstance(row[0], str) and 'Empr.:' in row[0]:
            if current_employee:
                employees_data.append(current_employee)
            current_employee = {
                'Código': str(row[5]) if row[5] is not None else None,  # Coluna F
                'Nome': ' '.join(str(cell) for cell in row[9:] if cell is not None),  # Coluna J em diante
                'Vínculo': None,
                'Cargo': None,
                'Salário': None,
                'Líquido': None
            }
        elif current_employee:
            if row[0] and isinstance(row[0], str):
                if 'Vínculo:' in row[0]:
                    current_employee['Vínculo'] = row[9] if row[9] is not None else None  # Coluna J
                elif 'Cargo:' in row[0]:
                    current_employee['Cargo'] = ' '.join(str(cell) for cell in row[9:] if cell is not None)  # Coluna J em diante
            
            # Salário na coluna BQ (69ª coluna) e Líquido na coluna BU (73ª coluna)
            current_employee['Salário'] = row[69] if len(row) > 69 and row[69] is not None else current_employee['Salário']
            current_employee['Líquido'] = row[73] if len(row) > 73 and row[73] is not None else current_employee['Líquido']

    if current_employee:
        employees_data.append(current_employee)

    return employees_data

def find_all_employees(sheet):
    employees = []
    for row in sheet.iter_rows(min_row=2):
        empr_cell = row[0]
        if isinstance(empr_cell.value, str) and 'Empr.' in empr_cell.value:
            codigo = sheet.cell(row=empr_cell.row, column=6).value  # Coluna F
            nome = sheet.cell(row=empr_cell.row, column=10).value  # Coluna J
            if codigo is not None and nome is not None:
                employees.append((str(codigo), str(nome)))
    return employees

# Caminho completo para o arquivo Excel
file_path = 'C:/Docs/Extrato Mensal.xlsx'

try:
    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Lê todos os colaboradores e códigos
    all_employees = find_all_employees(sheet)

    # Imprime todos os colaboradores e códigos
    print("\nTodos os Colaboradores:")
    for codigo, nome in all_employees:
        print(f"Código: {codigo}, Nome: {nome}")

    # Lê todos os dados detalhados dos colaboradores
    all_employees_data = find_all_employees_data(sheet)

    # Imprime os dados detalhados de todos os colaboradores
    print("\nDados detalhados de todos os Colaboradores:")
    for i, employee in enumerate(all_employees_data, 1):
        print(f"\n{'=' * 50}")
        print(f"Colaborador {i}:")
        print(f"{'=' * 50}")
        print(f"{'Código:':<10} {employee['Código']}")
        print(f"{'Nome:':<10} {employee['Nome']}")
        print(f"{'Vínculo:':<10} {employee['Vínculo']}")
        print(f"{'Cargo:':<10} {employee['Cargo']}")
        
        salario = employee['Salário']
        if salario is not None:
            try:
                salario_float = float(str(salario).replace(',', '.'))
                print(f"{'Salário:':<10} R$ {salario_float:,.2f}")
            except ValueError:
                print(f"{'Salário:':<10} {salario}")
        else:
            print(f"{'Salário:':<10} Não informado")
        
        liquido = employee['Líquido']
        if liquido is not None:
            try:
                liquido_float = float(str(liquido).replace(',', '.'))
                print(f"{'Líquido:':<10} R$ {liquido_float:,.2f}")
            except ValueError:
                print(f"{'Líquido:':<10} {liquido}")
        else:
            print(f"{'Líquido:':<10} Não informado")

    # Fecha o arquivo Excel
    workbook.close()

except FileNotFoundError:
    print(f"Erro: O arquivo '{file_path}' não foi encontrado.")
except Exception as e:
    print(f"Erro ao processar o arquivo: {e}")
