# Aqui está o script completo, agora incluindo a funcionalidade para calcular e mostrar os totais de salários e o total de colaboradores, além de garantir que o arquivo "Resultado.xlsx" seja gerado corretamente no diretório especificado. O script também imprimirá todos os dados detalhados dos colaboradores no terminal antes de salvar.

# ```python
# import openpyxl
# import pandas as pd
# import os

# def find_all_employees_data(sheet):
#     employees_data = []
#     current_employee = None

#     for row in sheet.iter_rows(min_row=1, values_only=True):
#         if row[0] and isinstance(row[0], str) and 'Empr.:' in row[0]:
#             if current_employee:
#                 employees_data.append(current_employee)
#             current_employee = {
#                 'Código': str(row[5]) if row[5] is not None else None,  # Coluna F
#                 'Nome': ' '.join(str(cell) for cell in row[9:] if cell is not None),  # Coluna J em diante
#                 'Vínculo': None,
#                 'Cargo': None,
#                 'Salário': None,
#                 'Líquido': None
#             }
#         elif current_employee:
#             if row[0] and isinstance(row[0], str):
#                 if 'Vínculo:' in row[0]:
#                     current_employee['Vínculo'] = row[9] if row[9] is not None else None  # Coluna J
#                 elif 'Cargo:' in row[0]:
#                     current_employee['Cargo'] = ' '.join(str(cell) for cell in row[9:] if cell is not None)  # Coluna J em diante
            
#             # Salário na coluna BQ (69) e Líquido na coluna BU (73)
#             current_employee['Salário'] = row[68] if len(row) > 68 and row[68] is not None else current_employee['Salário']
#             current_employee['Líquido'] = row[72] if len(row) > 72 and row[72] is not None else current_employee['Líquido']

#     if current_employee:
#         employees_data.append(current_employee)

#     return employees_data

# def find_all_employees(sheet):
#     employees = []
#     for row in sheet.iter_rows(min_row=2):
#         empr_cell = row[0]
#         if isinstance(empr_cell.value, str) and 'Empr.' in empr_cell.value:
#             codigo = sheet.cell(row=empr_cell.row, column=6).value  # Coluna F
#             nome = sheet.cell(row=empr_cell.row, column=10).value  # Coluna J
#             if codigo is not None and nome is not None:
#                 employees.append((str(codigo), str(nome)))
#     return employees

# # Caminho completo para o arquivo Excel
# file_path = r'C:\Docs\Extrato Mensal.xlsx'

# try:
#     # Carrega o arquivo Excel
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active

#     # Lê todos os colaboradores e códigos
#     all_employees = find_all_employees(sheet)

#     # Imprime todos os colaboradores e códigos
#     print("\nTodos os Colaboradores:")
#     for codigo, nome in all_employees:
#         print(f"Código: {codigo}, Nome: {nome}")

#     # Lê todos os dados detalhados dos colaboradores
#     all_employees_data = find_all_employees_data(sheet)

#     # Imprime os dados detalhados de todos os colaboradores
#     print("\nDados detalhados de todos os Colaboradores:")
#     for i, employee in enumerate(all_employees_data, 1):
#         print(f"\n{'=' * 50}")
#         print(f"Colaborador {i}:")
#         print(f"{'=' * 50}")
#         print(f"{'Código:':<10} {employee['Código']}")
#         print(f"{'Nome:':<10} {employee['Nome']}")
#         print(f"{'Vínculo:':<10} {employee['Vínculo']}")
#         print(f"{'Cargo:':<10} {employee['Cargo']}")
        
#         salario = employee['Salário']
#         if salario is not None:
#             try:
#                 salario_float = float(str(salario).replace(',', '.'))
#                 print(f"{'Salário:':<10} R$ {salario_float:,.2f}")
#             except ValueError:
#                 print(f"{'Salário:':<10} {salario}")
#         else:
#             print(f"{'Salário:':<10} Não informado")
        
#         liquido = employee['Líquido']
#         if liquido is not None:
#             try:
#                 liquido_float = float(str(liquido).replace(',', '.'))
#                 print(f"{'Líquido:':<10} R$ {liquido_float:,.2f}")
#             except ValueError:
#                 print(f"{'Líquido:':<10} {liquido}")
#         else:
#             print(f"{'Líquido:':<10} Não informado")

#     # Calcula os totais e contagens
#     total_salarios_celetistas = sum(float(emp['Salário'].replace(',', '.')) if isinstance(emp['Salário'], str) else (emp['Salário'] or 0) for emp in all_employees_data if emp['Vínculo'] == 'Celetista' and emp['Salário'] is not None)
#     total_salarios_aprendiz = sum(float(emp['Salário'].replace(',', '.')) if isinstance(emp['Salário'], str) else (emp['Salário'] or 0) for emp in all_employees_data if emp['Vínculo'] == 'Aprendiz' and emp['Salário'] is not None)
#     total_liquido = sum(float(emp['Líquido'].replace(',', '.')) if isinstance(emp['Líquido'], str) else (emp['Líquido'] or 0) for emp in all_employees_data if emp['Líquido'] is not None)
#     total_colaboradores = len(all_employees_data)
#     num_celetistas = sum(1 for emp in all_employees_data if emp['Vínculo'] == 'Celetista')
#     num_aprendizes = sum(1 for emp in all_employees_data if emp['Vínculo'] == 'Aprendiz')

#     # Imprime os totais e contagens no terminal
#     print(f"\n{'=' * 50}")
#     print("Totalizadores:")
#     print(f"{'=' * 50}")
#     print(f"Total de Salários para Celetistas: R$ {total_salarios_celetistas:,.2f}")
#     print(f"Total de Salários para Aprendizes: R$ {total_salarios_aprendiz:,.2f}")
#     print(f"Total Líquido: R$ {total_liquido:,.2f}")
#     print(f"Total de Colaboradores: {total_colaboradores}")
#     print(f"Número de Celetistas: {num_celetistas}")
#     print(f"Número de Aprendizes: {num_aprendizes}")

#     # Cria um DataFrame com os dados dos colaboradores
#     df = pd.DataFrame(all_employees_data)

#     # Adiciona uma linha com os totalizadores ao DataFrame
#     totalizadores = pd.DataFrame({
#         'Código': ['Total'],
#         'Nome': [''],
#         'Vínculo': [''],
#         'Cargo': [''],
#         'Salário': [f"R$ {total_salarios_celetistas + total_salarios_aprendiz:,.2f}"],
#         'Líquido': [f"R$ {total_liquido:,.2f}"]
#     })
#     df = pd.concat([df, totalizadores], ignore_index=True)

#     # Define o caminho para o arquivo de saída
#     output_path = r'C:\Docs\Resultado.xlsx'

#     # Salva o DataFrame como um arquivo Excel
#     with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='Dados dos Colaboradores')

#         # Ajusta a largura das colunas
#         for column in df:
#             column_length = max(df[column].astype(str).map(len).max(), len(column))
#             col_idx = df.columns.get_loc(column)
#             writer.sheets['Dados dos Colaboradores'].column_dimensions[openpyxl.utils.get_column_letter(col_idx+1)].width = column_length + 2

#     print(f"Os dados foram salvos com sucesso em: {output_path}")

#     # Fecha o arquivo Excel original
#     workbook.close()

# except FileNotFoundError:
#     print(f"Erro: O arquivo '{file_path}' não foi encontrado.")
# except Exception as e:
#     print(f"Erro ao processar o arquivo: {e}")
# ```

# ### Principais Funcionalidades do Script:

# 1. **Leitura e Extração de Dados**:
#    - O script lê os dados dos colaboradores da planilha, incluindo Código, Nome, Vínculo, Cargo, Salário e Líquido.

# 2. **Cálculo de Totais**:
#    - Calcula o total de salários para Celetistas e Aprendizes.
#    - Calcula o total líquido.
#    - Conta o total de colaboradores e o número de Celetistas e Aprendizes.

# 3. **Impressão dos Resultados**:
#    - Imprime todos os dados detalhados dos colaboradores no terminal.
#    - Exibe os totais e contagens no terminal.

# 4. **Criação e Salvamento do Arquivo Excel**:
#    - Cria um DataFrame com os dados dos colaboradores e adiciona uma linha com os totalizadores.
#    - Salva o DataFrame em um arquivo Excel chamado "Resultado.xlsx" no diretório especificado.
#    - Ajusta automaticamente a largura das colunas no arquivo de saída.

# Certifique-se de que o caminho do arquivo (`file_path`) está correto para o seu sistema. Se você encontrar algum problema ou precisar de mais ajustes, por favor, me informe.

# Citations:
# [1] https://aws.amazon.com/pt/what-is/python/
# [2] https://www.youtube.com/watch?v=A6cvOReAReM
# [3] https://www.dataquest.io/blog/reading-excel-file-python/
# [4] https://pt.wikipedia.org/wiki/Python
# [5] https://learn.microsoft.com/pt-br/sql/machine-learning/tutorials/python-taxi-classification-explore-data?view=sql-server-ver16

# Totalizadores:
# ==================================================
# Total de Sal�rios para Celetistas: R$ 574,817.95
# Total de Sal�rios para Aprendizes: R$ 3,400.00
# Total L�quido: R$ 285,624.94
# Total de Colaboradores: 245
# N�mero de Celetistas: 241
# N�mero de Aprendizes: 4
# Os dados foram salvos com sucesso em: C:\Docs\Resultado.xlsx

# [Done] exited with code=0 in 8.072 seconds

import openpyxl
import pandas as pd
import os

def find_all_employees_data(sheet):
    employees_data = []
    current_employee = None

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[0] and isinstance(row[0], str) and 'Empr.:' in row[0]:
            if current_employee:
                employees_data.append(current_employee)
            current_employee = {
                'Código': str(row[5]) if row[5] is not None else None,  # Coluna F
                'Nome': ' '.join(str(cell) for cell in row[9:] if cell is not None),  # Coluna J em diante
                'Vínculo': None,
                'Cargo': None,
                'Salário': None,
                'Líquido': None
            }
        elif current_employee:
            if row[0] and isinstance(row[0], str):
                if 'Vínculo:' in row[0]:
                    current_employee['Vínculo'] = row[9] if row[9] is not None else None  # Coluna J
                elif 'Cargo:' in row[0]:
                    current_employee['Cargo'] = ' '.join(str(cell) for cell in row[9:] if cell is not None)  # Coluna J em diante
            
            # Salário na coluna BQ (69) e Líquido na coluna BU (73)
            current_employee['Salário'] = row[68] if len(row) > 68 and row[68] is not None else current_employee['Salário']
            current_employee['Líquido'] = row[72] if len(row) > 72 and row[72] is not None else current_employee['Líquido']

    if current_employee:
        employees_data.append(current_employee)

    return employees_data

def find_all_employees(sheet):
    employees = []
    for row in sheet.iter_rows(min_row=2):
        empr_cell = row[0]
        if isinstance(empr_cell.value, str) and 'Empr.' in empr_cell.value:
            codigo = sheet.cell(row=empr_cell.row, column=6).value  # Coluna F
            nome = sheet.cell(row=empr_cell.row, column=10).value  # Coluna J
            if codigo is not None and nome is not None:
                employees.append((str(codigo), str(nome)))
    return employees

# Caminho completo para o arquivo Excel
file_path = r'C:\Docs\Extrato Mensal.xlsx'

try:
    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Lê todos os colaboradores e códigos
    all_employees = find_all_employees(sheet)

    # Imprime todos os colaboradores e códigos
    print("\nTodos os Colaboradores:")
    for codigo, nome in all_employees:
        print(f"Código: {codigo}, Nome: {nome}")

    # Lê todos os dados detalhados dos colaboradores
    all_employees_data = find_all_employees_data(sheet)

    # Imprime os dados detalhados de todos os colaboradores
    print("\nDados detalhados de todos os Colaboradores:")
    for i, employee in enumerate(all_employees_data, 1):
        print(f"\n{'=' * 50}")
        print(f"Colaborador {i}:")
        print(f"{'=' * 50}")
        print(f"{'Código:':<10} {employee['Código']}")
        print(f"{'Nome:':<10} {employee['Nome']}")
        print(f"{'Vínculo:':<10} {employee['Vínculo']}")
        print(f"{'Cargo:':<10} {employee['Cargo']}")
        
        salario = employee['Salário']
        if salario is not None:
            try:
                salario_float = float(str(salario).replace(',', '.'))
                print(f"{'Salário:':<10} R$ {salario_float:,.2f}")
            except ValueError:
                print(f"{'Salário:':<10} {salario}")
        else:
            print(f"{'Salário:':<10} Não informado")
        
        liquido = employee['Líquido']
        if liquido is not None:
            try:
                liquido_float = float(str(liquido).replace(',', '.'))
                print(f"{'Líquido:':<10} R$ {liquido_float:,.2f}")
            except ValueError:
                print(f"{'Líquido:':<10} {liquido}")
        else:
            print(f"{'Líquido:':<10} Não informado")

    # Calcula os totais e contagens
    total_salarios_celetistas = sum(float(emp['Salário'].replace(',', '.')) if isinstance(emp['Salário'], str) else (emp['Salário'] or 0) for emp in all_employees_data if emp['Vínculo'] == 'Celetista' and emp['Salário'] is not None)
    total_salarios_aprendiz = sum(float(emp['Salário'].replace(',', '.')) if isinstance(emp['Salário'], str) else (emp['Salário'] or 0) for emp in all_employees_data if emp['Vínculo'] == 'Aprendiz' and emp['Salário'] is not None)
    total_liquido = sum(float(emp['Líquido'].replace(',', '.')) if isinstance(emp['Líquido'], str) else (emp['Líquido'] or 0) for emp in all_employees_data if emp['Líquido'] is not None)
    total_colaboradores = len(all_employees_data)
    num_celetistas = sum(1 for emp in all_employees_data if emp['Vínculo'] == 'Celetista')
    num_aprendizes = sum(1 for emp in all_employees_data if emp['Vínculo'] == 'Aprendiz')

    # Imprime os totais e contagens no terminal
    print(f"\n{'=' * 50}")
    print("Totalizadores:")
    print(f"{'=' * 50}")
    print(f"Total de Salários para Celetistas: R$ {total_salarios_celetistas:,.2f}")
    print(f"Total de Salários para Aprendizes: R$ {total_salarios_aprendiz:,.2f}")
    print(f"Total Líquido: R$ {total_liquido:,.2f}")
    print(f"Total de Colaboradores: {total_colaboradores}")
    print(f"Número de Celetistas: {num_celetistas}")
    print(f"Número de Aprendizes: {num_aprendizes}")

    # Cria um DataFrame com os dados dos colaboradores
    df = pd.DataFrame(all_employees_data)

    # Adiciona uma linha com os totalizadores ao DataFrame
    totalizadores = pd.DataFrame({
        'Código': ['Total'],
        'Nome': [''],
        'Vínculo': [''],
        'Cargo': [''],
        'Salário': [f"R$ {total_salarios_celetistas + total_salarios_aprendiz:,.2f}"],
        'Líquido': [f"R$ {total_liquido:,.2f}"]
    })
    df = pd.concat([df, totalizadores], ignore_index=True)

    # Define o caminho para o arquivo de saída
    output_path = r'C:\Docs\Resultado.xlsx'

    # Salva o DataFrame como um arquivo Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados dos Colaboradores')

        # Ajusta a largura das colunas
        for column in df:
            column_length = max(df[column].astype(str).map(len).max(), len(column))
            col_idx = df.columns.get_loc(column)
            writer.sheets['Dados dos Colaboradores'].column_dimensions[openpyxl.utils.get_column_letter(col_idx+1)].width = column_length + 2

    print(f"Os dados foram salvos com sucesso em: {output_path}")

    # Fecha o arquivo Excel original
    workbook.close()

except FileNotFoundError:
    print(f"Erro: O arquivo '{file_path}' não foi encontrado.")
except Exception as e:
    print(f"Erro ao processar o arquivo: {e}")
