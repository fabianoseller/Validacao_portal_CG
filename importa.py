# # -----------------------------------------------------------------
# # python cg
# # https://hub.asimov.academy/tutorial/lendo-e-escrevendo-planilhas-com-python/

# # https://medium.com/@andersoneduardo_50576/abrindo-arquivos-excel-em-python-833de325df70

# # https://hub.asimov.academy/blog/como-automatizar-seu-excel-com-python/

# # https://www.hashtagtreinamentos.com/integrar-python-com-excel-python

# # https://acervolima.com/como-extrair-dados-do-arquivo-excel-usando-o-pandas/

# # https://pt.stackoverflow.com/questions/426833/como-ler-uma-tabela-de-excel-no-pandas-pulando-as-primeiras-linhas-sem-perder-i

# # https://pt.stackoverflow.com/questions/97414/ler-linhas-e-colunas-pelo-python-pelo-excel

# # https://pt.stackoverflow.com/questions/142118/ler-informa%C3%A7%C3%A3o-de-excel-em-forma-de-matriz-com-python
# # https://www.freecodecamp.org/portuguese/news/como-criar-ler-atualizar-e-pesquisar-atraves-de-arquivos-do-excel-usando-o-python/


# # html form depois post

# # wb = load_workbook("/caminho/completo/para/Resultados.xlsx")
# # Caminho Relativo: Se o arquivo “Resultados.xlsx” estiver no mesmo diretório (pasta) que o seu script Python, você pode simplesmente usar o nome do arquivo:
# # Python

# # wb = load_workbook("Resultados.xlsx")
# # Código gerado por IA. Examine e use com cuidado. Mais informações em perguntas frequentes.
# # Caminho Absoluto: Se o arquivo estiver em um diretório diferente, especifique o caminho completo para o arquivo. Por exemplo:
# # Python

# wb = load_workbook("/caminho/completo/para/Resultados.xlsx")

# from openpyxl import load_workbook

# # Carrega o arquivo Excel
# wb = load_workbook("Resultados.xlsx")
# teste


# # Define o número de células com dados
# Nod = 49

# # Seleciona a planilha desejada (no seu caso, 'Primeira')
# sheet1 = wb['Primeira']

# # Inicializa uma lista para armazenar os valores de pH
# ph_value = []

# # Itera pelas colunas (de 1 a Nod)
# for a in range(1, Nod + 1):
#     # Lê o valor da célula na linha 3 e coluna 'a'
#     ph_value.append(sheet1.cell(row=3, column=a).value)

# # Agora 'ph_value' contém os valores de pH das células especificadas
# print(ph_value)
# ----------------------------------------

# from openpyxl import load_workbook

# # Carrega o arquivo Excel
# wb = load_workbook("Resultados.xlsx")

# # Define o número de células com dados
# Nod = 49

# # Seleciona a planilha desejada (no seu caso, 'Primeira')
# sheet1 = wb['Primeira']

# # Inicializa uma lista para armazenar os valores de pH
# ph_values = []

# # Itera pelas colunas (de 1 a Nod)
# for col in range(1, Nod + 1):
#     # Inicializa uma lista para armazenar os valores da coluna atual
#     col_values = []
#     for row in range(1, sheet1.max_row + 1):
#         # Lê o valor da célula na linha 'row' e coluna 'col'
#         col_values.append(sheet1.cell(row=row, column=col).value)
#     # Adiciona os valores da coluna à lista principal
#     ph_values.append(col_values)

# # Agora 'ph_values' contém todos os valores de pH das células especificadas
# # Cada sublista representa uma coluna
# print(ph_values)

import openpyxl
import os

# Verifica o diretório atual de trabalho
print("Diretório atual:", os.getcwd())

# Caminho completo para o arquivo Excel
file_path = 'C:/Docs/Extrato Mensal.xlsx'

# Carrega o arquivo Excel
try:
    theFile = openpyxl.load_workbook(file_path)
    print("Planilhas disponíveis:", theFile.sheetnames)

    # Seleciona a primeira planilha
    sheet = theFile.active

    # Obtém as dimensões da planilha
    max_row = sheet.max_row
    max_column = sheet.max_column

    print(f"A planilha tem {max_row} linhas e {max_column} colunas.")

    # Lê os dados da planilha
    data = []
    for row in range(1, max_row + 1):  # Lê todas as linhas
        row_data = []
        for col in range(1, max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        data.append(row_data)

    # Imprime os primeiros 5 registros como exemplo
    print("\nPrimeiros 5 registros:")
    for i, row in enumerate(data[:5], start=1):
        print(f"Registro {i}:", row)

    # Fecha o arquivo Excel
    theFile.close()

except FileNotFoundError:
    print(f"Erro: O arquivo '{file_path}' não foi encontrado.")
except Exception as e:
    print(f"Erro ao processar o arquivo: {e}")