import openpyxl
import re
from datetime import datetime

def find_value(sheet, field):
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and field in cell.value:
                if field == 'Empr.':
                    # Extrai o código (coluna F, que é a 6ª coluna)
                    codigo = sheet.cell(row=cell.row, column=6).value
                    # Extrai o nome (sempre na coluna J)
                    nome = sheet.cell(row=cell.row, column=10).value  # Coluna J é a 10ª coluna
                    return (str(codigo) if codigo is not None else None, 
                            str(nome) if nome is not None else None)
                elif field in ['Cargo:', 'Salário:', 'Líquido:']:
                    # Retorna o segundo valor após o campo
                    valor = sheet.cell(row=cell.row, column=cell.column+2).value
                    return str(valor) if valor is not None else None
                else:
                    # Retorna o valor após o campo
                    valor = sheet.cell(row=cell.row, column=cell.column+1).value
                    return str(valor) if valor is not None else None
    return None

# Caminho completo para o arquivo Excel
file_path = 'C:/Docs/Extrato Mensal.xlsx'

try:
    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Campos a serem extraídos
    fields = {
        'Empr.': ('Código', 'Colaborador'),
        'Vínculo:': 'Vínculo',
        'Cargo:': 'Cargo',
        'Salário:': 'Salário',
        'Líquido:': 'Líquido',
        'Situação:': 'Situação',
        'CPF:': 'CPF',
        'Adm:': 'Data de Admissão',
        'CC:': 'Centro de Custo',
        'Depto:': 'Departamento',
        'Horas Mês:': 'Horas Mês',
        'C.B.O:': 'CBO',
        'Filial:': 'Filial',
        'Proventos:': 'Proventos',
        'Descontos:': 'Descontos'
    }

    # Dicionário para armazenar os dados extraídos
    extracted_data = {}

    # Extrai os dados
    for field, key in fields.items():
        value = find_value(sheet, field)
        if value:
            if isinstance(value, tuple):
                extracted_data['Código'], extracted_data['Colaborador'] = value
            else:
                # Remove espaços em branco extras e caracteres especiais
                if isinstance(value, str):
                    value = re.sub(r'\s+', ' ', value).strip()
                extracted_data[key] = value

    # Imprime os dados extraídos
    print("\nDados extraídos:")
    for key, value in extracted_data.items():
        print(f"{key}: {value}")

    # Fecha o arquivo Excel
    workbook.close()

except FileNotFoundError:
    print(f"Erro: O arquivo '{file_path}' não foi encontrado.")
except Exception as e:
    print(f"Erro ao processar o arquivo: {e}")
