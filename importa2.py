import openpyxl
import os
from datetime import datetime

# Verifica o diretório atual de trabalho
print("Diretório atual:", os.getcwd())

# Caminho completo para o arquivo Excel
file_path = 'C:/Docs/Extrato Mensal.xlsx'

# Campos a serem extraídos
campos_desejados = ['Cód', 'Nome', 'Vinculo', 'Cargo', 'Salario Contratado', 'Salario Liquido']

# Carrega o arquivo Excel
try:
    theFile = openpyxl.load_workbook(file_path)
    print("Planilhas disponíveis:", theFile.sheetnames)

    # Seleciona a planilha específica chamada "Extrato Mensal form"
    if "Extrato Mensal form" in theFile.sheetnames:
        sheet = theFile["Extrato Mensal form"]
    else:
        raise ValueError("A planilha 'Extrato Mensal form' não foi encontrada.")

    # Obtém as dimensões da planilha
    max_row = sheet.max_row
    max_column = sheet.max_column

    print(f"A planilha tem {max_row} linhas e {max_column} colunas.")

    # Encontra os índices das colunas desejadas
    headers = []
    column_indices = {}
    for col in range(1, max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        headers.append(cell_value)
        if cell_value in campos_desejados:
            column_indices[cell_value] = col

    # Verifica se todos os campos desejados foram encontrados
    if len(column_indices) != len(campos_desejados):
        missing_fields = set(campos_desejados) - set(column_indices.keys())
        print(f"Aviso: Os seguintes campos não foram encontrados: {missing_fields}")

    # Lê os dados da planilha
    data = []
    for row in range(2, max_row + 1):  # Começa da segunda linha para pular os cabeçalhos
        row_data = {}
        for campo in campos_desejados:
            if campo in column_indices:
                cell_value = sheet.cell(row=row, column=column_indices[campo]).value
                row_data[campo] = cell_value
        data.append(row_data)

    # Imprime todos os registros
    print("\nRegistros extraídos:")
    for i, row in enumerate(data, start=1):
        print(f"Registro {i}:", row)

    # Fecha o arquivo Excel
    theFile.close()

except FileNotFoundError:
    print(f"Erro: O arquivo '{file_path}' não foi encontrado.")
except ValueError as ve:
    print(ve)
except Exception as e:
    print(f"Erro ao processar o arquivo: {e}")