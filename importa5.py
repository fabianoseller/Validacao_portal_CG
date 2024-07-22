import openpyxl
import re

def find_all_employees_data(sheet):
    employees_data = []
    for row in sheet.iter_rows(min_row=2):  # Começamos da segunda linha, assumindo que a primeira é cabeçalho
        empr_cell = row[0]  # Assumindo que 'Empr.' está sempre na primeira coluna
        if isinstance(empr_cell.value, str) and 'Empr.' in empr_cell.value:
            codigo = sheet.cell(row=empr_cell.row, column=6).value  # Coluna F
            nome = sheet.cell(row=empr_cell.row, column=10).value  # Coluna J
            vinculo = find_value_in_column(sheet, empr_cell.row, 'Vínculo:', 10)  # Coluna J
            cargo = find_value_in_column(sheet, empr_cell.row, 'Cargo:', 10)  # Coluna J
            salario = find_value_in_row(sheet, empr_cell.row, 'Salário:')
            liquido = find_value_in_row(sheet, empr_cell.row, 'Líquido:')
            
            employees_data.append({
                'Código': str(codigo) if codigo is not None else None,
                'Nome': str(nome) if nome is not None else None,
                'Vínculo': vinculo,
                'Cargo': cargo,
                'Salário': salario,
                'Líquido': liquido
            })
    return employees_data

def find_value_in_row(sheet, row, field):
    for cell in sheet[row]:
        if isinstance(cell.value, str) and field in cell.value:
            if field in ['Cargo:', 'Salário:', 'Líquido:']:
                valor = sheet.cell(row=row, column=cell.column+2).value
            else:
                valor = sheet.cell(row=row, column=cell.column+1).value
            return str(valor) if valor is not None else None
    return None

def find_value_in_column(sheet, start_row, field, column):
    for row in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column).value
        if isinstance(cell_value, str) and field in cell_value:
            return sheet.cell(row=row, column=column+1).value
    return None

def find_all_employees(sheet):
    employees = []
    for row in sheet.iter_rows(min_row=2):
        empr_cell = row[0]
        if isinstance(empr_cell.value, str) and 'Empr.' in empr_cell.value:
            codigo = sheet.cell(row=empr_cell.row, column=6).value  # Coluna F
            nome = sheet.cell(row=empr_cell.row, column=10).value  # Coluna J
            if codigo is not None and nome is not None:
                employees.append((str(codigo), str(nome)))
    return employees

def find_value(sheet, field):
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and field in cell.value:
                if field == 'Empr.':
                    codigo = sheet.cell(row=cell.row, column=6).value
                    nome = sheet.cell(row=cell.row, column=10).value
                    return (str(codigo) if codigo is not None else None, 
                            str(nome) if nome is not None else None)
                elif field in ['Cargo:', 'Salário:', 'Líquido:']:
                    valor = sheet.cell(row=cell.row, column=cell.column+2).value
                    return str(valor) if valor is not None else None
                else:
                    valor = sheet.cell(row=cell.row, column=cell.column+1).value
                    return str(valor) if valor is not None else None
    return None

# Caminho completo para o arquivo Excel
file_path = 'C:/Docs/Extrato Mensal.xlsx'

try:
    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Lê todos os colaboradores e códigos
    all_employees = find_all_employees(sheet)

    # Imprime todos os colaboradores e códigos
    print("\nTodos os Colaboradores:")
    for codigo, nome in all_employees:
        print(f"Código: {codigo}, Nome: {nome}")

    # Lê todos os dados detalhados dos colaboradores
    all_employees_data = find_all_employees_data(sheet)

    # Imprime os dados detalhados de todos os colaboradores
    print("\nDados detalhados de todos os Colaboradores:")
    for employee in all_employees_data:
        print("\n" + "="*50)
        for key, value in employee.items():
            print(f"{key}: {value}")

    # Fecha o arquivo Excel
    workbook.close()

except FileNotFoundError:
    print(f"Erro: O arquivo '{file_path}' não foi encontrado.")
except Exception as e:
    print(f"Erro ao processar o arquivo: {e}")
